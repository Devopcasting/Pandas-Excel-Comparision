import os
from fastapi import APIRouter, HTTPException, Request
from fastapi.templating import Jinja2Templates
from urllib.parse import unquote
from openpyxl import load_workbook
import pandas as pd

excel_router = APIRouter()
templates = Jinja2Templates(directory=r"app\templates")


# Check if sheet is blank
def is_sheet_blank(file_path: str, sheet_number: int) -> bool:
    # Load the workbook
    workbook = load_workbook(file_path)
    # Get the sheet by index
    sheet = workbook.worksheets[sheet_number - 1]

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                return False  # If any cell has a value, the sheet is not blank
    return True  # If no cell has a value, the sheet is blank

# Check for valid sheet number
def is_sheet_invalid(file_path: str, sheet_number: int) -> bool:
    # Load the workbook
    workbook = load_workbook(file_path)
    # Check if the provided sheet number is valid
    if sheet_number < 1 or sheet_number > len(workbook.sheetnames):
        return False
    return True

@excel_router.get("/compare_excel")
async def compare_excel(request: Request, file_url1: str, file_url2: str, file1_sheet_number: int=1, file2_sheet_number: int=1):
    # Reset the File path string
    path_file1 = unquote(file_url1)
    path_file2 = unquote(file_url2)

    # Check if the Excel documents are available in the path
    if not os.path.exists(path_file1.replace("%20", " ")):
        raise HTTPException(status_code=404, detail=f"File {path_file1} not found")
    if not os.path.exists(path_file2.replace("%20", " ")):
        raise HTTPException(status_code=404, detail=f"File {path_file2} not found")
    
    # Check if both files are valid xlsx format files
    if not path_file1.endswith('.xlsx') or not path_file2.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Files must be in XLSX format")
    
    # Check if both files are valid XLSX format files
    try:
        load_workbook(path_file1)
        load_workbook(path_file2)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file format: {str(e)}")
    
    # Check if the sheet number is valid
    if not is_sheet_invalid(path_file1, file1_sheet_number):
        raise HTTPException(status_code=404, detail=f"File {path_file1} Invalid sheet number {file1_sheet_number}")
    if not is_sheet_invalid(path_file2, file2_sheet_number):
        raise HTTPException(status_code=404, detail=f"File {path_file2} Invalid sheet number {file2_sheet_number}")
    
    # Check if the Sheet is not blank
    if is_sheet_blank(path_file1, file1_sheet_number):
        raise HTTPException(status_code=404, detail=f"File {path_file1} sheet number {file1_sheet_number} is empty")
    if is_sheet_blank(path_file2, file2_sheet_number):
        raise HTTPException(status_code=404, detail=f"File {path_file2} sheet number {file2_sheet_number} is empty")

    
    # Read the Excel files using Pandas
    df1 = pd.read_excel(path_file1, sheet_name=file1_sheet_number -1)
    df2 = pd.read_excel(path_file2, sheet_name=file2_sheet_number -1)

    # Ensure both the DataFrames have the same columns
    columns = list(set(df1.columns) | set(df2.columns))
    df1 = df1.reindex(columns=columns)
    df2 = df2.reindex(columns=columns)

    # Find the length of each DataFrame
    len_df1 = len(df1)
    len_df2 = len(df2)

    # Check which DataFrame has more rows
    if len_df1 > len_df2:
        # Append missing rows in df2 with None values
        missing_rows = len_df1 - len_df2
        df2 = pd.concat([df2, pd.DataFrame([[None]*len(columns)]*missing_rows, columns=columns)], ignore_index=True)
    elif len_df2 > len_df1:
        # Append missing rows in df1 with None values
        missing_rows = len_df2 - len_df1
        df1 = pd.concat([df1, pd.DataFrame([[None]*len(columns)]*missing_rows, columns=columns)], ignore_index=True)

    # Highlight Rows
    diff_df = df1.compare(df2)
    highlighted_rows = diff_df.index.tolist()

    # Get the indices of all rows that are different or present only in one dataframe
    highlighted_rows = list(set(diff_df.index).union(set(df1.index).symmetric_difference(set(df2.index))))

    # Convert dataframes to HTML tables
    table1 = df1.to_dict(orient='records')
    table2 = df2.to_dict(orient='records')

    # Example Path
    # http://localhost:8000/compare_excel/?file_url1=C%3A%5CUsers%5CPrashant Pokhriyal%5CDownloads%5CLab13%5Cfile1.xlsx&file_url2=C%3A%5CUsers%5CPrashant Pokhriyal%5CDownloads%5CLab13%5Cfile2.xlsx
    return templates.TemplateResponse("compare_excel.html", {"request": request, "data1": table1, "data2": table2, 
                                                             "highlighted_rows": highlighted_rows, 
                                                             "title": "Contentverse Document Comparision", 
                                                             "file1":path_file1, "file2": path_file2,"file1_sheet_number": file1_sheet_number,"file2_sheet_number": file2_sheet_number })